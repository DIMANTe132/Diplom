import os
import webbrowser
from datetime import date
from textwrap import wrap
import PySimpleGUI as sg
import folium
import geopandas
import matplotlib.pyplot as plt
import osmnx as ox
import pandas as pd
import requests
import xlwings as xw
from branca.element import MacroElement, Template
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

N_MUN = 52
dictionary = {'Количество магазинов': ['магазины'],
              'Число спортивных сооружений': ['спортивные сооружения - всего'],
              'Вывоз твердых коммунальных отходов, тыс. куб. м': [
                  'вывезено за год твердых коммунальных отходов (тыс. куб. м)'],
              'Площадь жилых помещений, тыс. кв. м': ['общая площадь жилых помещений', 'весь жилищный фонд'],
              'Протяженность дорог с твердым покрытием, км': ['с твердым покрытием'],
              'Доходы местного бюджета, тыс. руб.': ['доходы местного бюджета, фактически исполненные', 'всего'],
              'Расходы местного бюджета, тыс. руб.': ['расходы  местного бюджета, фактически исполненные', 'всего'],
              'Число умерших': ['число умерших'],
              'Число родившихся': ['число родившихся (без мертворожденных)'],
              'Число прибывших': ['число прибывших', 'всего', 'всего', 'миграция-всего'],
              'Число убывших': ['число выбывших', 'всего', 'всего', 'миграция-всего'],
              'Число лечебно-профилактичнеских организаций': ['число лечебно-профилактических организаций'],
              'Число общеобразовательных организаций на начало учебного года, ед': [
                  'число общеобразовательных организаций  на начало учебного года'],
              'Численность работников организаций культурно-досугового типа, ед': [
                  'численность работников организаций культурно-досугового типа с учетом обособленных подразделений ('
                  'филиалов)'],
              'Число жителей': ['все население', 'на 1 января']}


def find_parameter(code, arr):
    index = 0
    for a in arr:
        index = code.find(a, index)
        if index == -1:
            return -1
        index = code.find('god', index) + 5
        if code[index] != '<':
            value = code[index:code.find('<', index)]
            if value.isdigit():
                return int(value)
            else:
                return float(value)
    return -1


def write_data(work_row):
    ws_wings.range(f"{get_column_letter(local_max_col - 1) + str(work_row)}"
                   f",{get_column_letter(local_max_col) + str(work_row)}").formula = \
        ws_wings.range(f"{get_column_letter(local_max_col - 1) + str(work_row)}"
                       f",{get_column_letter(local_max_col - 1) + str(work_row)}").formula

    ws_wings.range(
        f"{get_column_letter(local_max_col + len(local_years) + 2) + str(work_row)}"
        f",{get_column_letter(local_max_col + len(local_years) + 3) + str(work_row)}").formula = \
        ws_wings.range(
            f"{get_column_letter(local_max_col + len(local_years) + 2) + str(work_row)}"
            f",{get_column_letter(local_max_col + len(local_years) + 2) + str(work_row)}").formula


def check_data(work_row):
    result = find_parameter(cityInformation.text.lower(),
                            dictionary[ws_pyxl[get_column_letter(2)
                                               + str(work_row)].value])
    if result == -1:
        ws_wings.range(get_column_letter(local_max_col) + str(work_row)).value = \
            ws_wings.range(
                get_column_letter(local_max_col - 1) + str(work_row)).value
    else:
        ws_wings.range(
            get_column_letter(local_max_col) + str(work_row)).value = result


def get_max_elements():
    years = []
    year_to_col = {}
    max_col = 0
    for col in range(1, ws_pyxl.max_column):
        year = ws_pyxl[get_column_letter(col) + '4'].value
        if type(year) is int and year not in years:
            years.append(year)
            year_to_col[year] = col
            max_col = col

    return years, year_to_col, max_col


def prepare_data():
    global df_all, id2intensity
    osmnx_row['display_name'] = municipalities[i]
    folium.Marker(
        location=[osmnx_row.centroid.y, osmnx_row.centroid.x],
        icon=folium.DivIcon(html=f"""
          <div style="color:#000000;background:#0000fff;width:20px;text-align:center;font-size:8pt;">{i + 1}</div>
        """),
    ).add_to(m)
    df_all = df_all.append(osmnx_row, ignore_index=True)
    id2intensity = id2intensity.append({'osm_id': int(osmnx_row.osm_id), 'intensity': data[i]},
                                       ignore_index=True)


wb_pyxl = load_workbook('МУНИЦИПАЛИТЕТЫ.xlsx', data_only=True)
ws_pyxl = wb_pyxl['1']
municipalities = []
parameters = []
parameter_to_row = {}

years_mun, year_to_col_mun, max_col_mun = get_max_elements()

for i in range(5, 32):
    if i == 16:
        continue
    parameters.append(ws_pyxl[get_column_letter(2) + str(i)].value)
    parameter_to_row[ws_pyxl[get_column_letter(2) + str(i)].value] = i

for i in range(1, N_MUN + 1):
    ws_pyxl = wb_pyxl[str(i)]
    municipalities.append(ws_pyxl[get_column_letter(1) + '1'].value)

layout = [[sg.Text("Проверить обновление данных:")],
          [sg.Button("Актуализировать данные"), sg.Text(key='out1')],
          [sg.HorizontalSeparator()],
          [sg.Text("Построить график:")],
          [sg.Text("Муниципалитет:"), sg.Combo(municipalities, key='mun')],
          [sg.Text("Параметр:"), sg.Combo(parameters, key='par1')],
          [sg.Button("Построить график"), sg.Text(key='out2')],
          [sg.HorizontalSeparator()],
          [sg.Text("Показать карту со статистикой:")],
          [sg.Text("Год:"), sg.Combo(years_mun, key='years')],
          [sg.Text("Параметр:"), sg.Combo(parameters, key='par2')],
          [sg.Button("Визуализация"), sg.Text(key='out3')]]

window = sg.Window("Системы мониторинга экономического развития муниципалитетов НО ", layout)

while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED:
        break
    elif event == "Актуализировать данные":
        current_year = date.today().year
        if current_year in years_mun:
            window['out1'].update('Информация является актуальной!')
        else:
            for i in range(max(years_mun) + 1, current_year):
                year_to_col_mun[i] = year_to_col_mun[i - 1] + 1

            wb_wings = xw.Book('МУНИЦИПАЛИТЕТЫ.xlsx')
            f = requests.get("https://www.gks.ru/scripts/db_inet2/passport/munr.aspx?base=munst22")
            mainCode = f.text.lower()
            local_years = []
            local_max_col = 0
            for i in range(1, 39):
                ws_wings = wb_wings.sheets[str(i)]
                ws_pyxl = wb_pyxl[str(i)]
                local_years = years_mun.copy()
                local_max_col = max_col_mun
                for y in range(max(years_mun) + 1, current_year):
                    local_years.append(y)
                    local_max_col += 1
                    ws_wings.range(f'{get_column_letter(local_max_col)}:{get_column_letter(local_max_col)}').insert()
                    j = 33
                    while j < ws_pyxl.max_row:
                        current_row = j
                        city = ws_pyxl[get_column_letter(1) + str(current_row)].value
                        if city:
                            city = ':' + city.lower() + '@'
                            city_code = mainCode.find(city) - 8
                            cityInformation = requests.get(
                                f"https://www.gks.ru/scripts/db_inet2/passport/table.aspx?"
                                f"opt={mainCode[city_code:city_code + 8]}{y}")

                            current_row += 13
                            ws_wings.range(get_column_letter(local_max_col) + str(current_row)).value = y

                            for r in range(1, 16):
                                check_data(current_row + r)

                            current_row = j + 1
                            ws_wings.range(get_column_letter(local_max_col) + str(current_row)).value = \
                                ws_wings.range(get_column_letter(local_max_col + len(local_years) + 3) + str(
                                    current_row)).value = y
                            for r in range(1, 12):
                                write_data(current_row + r)

                        j += 30

                    for j in range(31, 3, -1):
                        if j == 4:
                            ws_wings.range(get_column_letter(local_max_col) + str(j)).value = \
                                ws_wings.range(
                                    get_column_letter(local_max_col + len(local_years) + 3) + str(j)).value = y
                        elif j == 16:
                            ws_wings.range(get_column_letter(local_max_col) + str(j)).value = y

                        else:
                            ws_wings.range(f"{get_column_letter(local_max_col - 1) + str(j)}"
                                           f",{get_column_letter(local_max_col) + str(j)}").formula = \
                                ws_wings.range(f"{get_column_letter(local_max_col - 1) + str(j)}"
                                               f",{get_column_letter(local_max_col - 1) + str(j)}").formula

                            if j < 16:
                                ws_wings.range(f"{get_column_letter(local_max_col + len(local_years) + 2) + str(j)}"
                                               f",{get_column_letter(local_max_col + len(local_years) + 3) + str(j)}").formula = \
                                    ws_wings.range(f"{get_column_letter(local_max_col + len(local_years) + 2) + str(j)}"
                                                   f",{get_column_letter(local_max_col + len(local_years) + 2) + str(j)}").formula

            years_mun = local_years
            max_col_mun = local_max_col

            ws_pyxl = wb_pyxl['39']
            years_city, year_to_col_city, max_col_city = get_max_elements()

            for i in range(39, N_MUN + 1):
                ws_wings = wb_wings.sheets[str(i)]
                ws_pyxl = wb_pyxl[str(i)]
                local_years = years_city.copy()
                local_max_col = max_col_city
                for y in range(max(years_city) + 1, current_year):
                    local_years.append(y)
                    local_max_col += 1
                    ws_wings.range(f'{get_column_letter(local_max_col)}:{get_column_letter(local_max_col)}').insert()

                    city = ws_pyxl[get_column_letter(1) + '1'].value
                    city = ' '.join(city.split()[2:4])
                    if city:
                        city = city.lower()
                        city_code = mainCode.find('\'>', mainCode.find(city) - 10)
                        cityInformation = requests.get(
                            f"https://www.gks.ru/scripts/db_inet2/passport/table.aspx?"
                            f"opt={mainCode[city_code:city_code + 8]}{y}")

                        ws_wings.range(get_column_letter(local_max_col) + '16').value = y

                        for r in range(17, 32):
                            check_data(r)

                        ws_wings.range(get_column_letter(local_max_col) + '4').value = \
                            ws_wings.range(get_column_letter(local_max_col + len(local_years) + 3) + '4').value = y
                        for r in range(5, 16):
                            write_data(r)

            wb_wings.save('qwe.xlsx')

    elif event == 'Построить график':
        if values['mun'] == '' or values['par1'] == '':
            window['out2'].update('Пожалуйста выберете муниципалитет и параметр!')
        else:
            window['out2'].update('')
            for i in range(1, N_MUN + 1):
                ws_pyxl = wb_pyxl[str(i)]
                if ws_pyxl[get_column_letter(1) + '1'].value == values['mun']:
                    break

            local_years = []
            for col in range(1, ws_pyxl.max_column):
                year = ws_pyxl[get_column_letter(col) + '4'].value
                if type(year) is int and year not in local_years:
                    local_years.append(year)

            row = 0
            for i in range(5, 32):
                if ws_pyxl[get_column_letter(2) + str(i)].value == values['par1']:
                    row = i
                    break

            data = []
            for i in range(3, 3 + len(local_years)):
                data.append(ws_pyxl[get_column_letter(i) + str(row)].value)

            plt.figure(figsize=(12, 6))
            plt.plot(local_years, data)
            plt.title("\n".join(wrap(f'График "{values["par1"].lower()}" в муниципалитете - {values["mun"]}', 60)))
            plt.xlabel('Year')
            if ',' in values['par1']:
                ed = values['par1'][values['par1'].find(',') + 2:].capitalize()
                plt.ylabel(ed)
            else:
                plt.ylabel('Шт.')
            plt.show()

    elif event == "Визуализация":
        if values['years'] == '' or values['par2'] == '':
            window['out3'].update('Пожалуйста выберете год и параметр!')
        else:
            data = []
            legend = []
            for i in range(1, N_MUN + 1):
                ws_pyxl = wb_pyxl[str(i)]
                if i < 39:
                    data.append(ws_pyxl[get_column_letter(year_to_col_mun[values['years']]) +
                                        str(parameter_to_row[values['par2']])].value)
                else:
                    data.append(ws_pyxl[get_column_letter(year_to_col_mun[values['years']] - 2) +
                                        str(parameter_to_row[values['par2']])].value)
                legend.append(str(i) + ' - ' + municipalities[i - 1])

            html_page = 'map.html'
            NO = ox.geocode_to_gdf('Нижегородская область')
            m = folium.Map()

            df_all = pd.DataFrame(NO, columns=NO.columns.values.tolist())
            id2intensity = pd.DataFrame(columns=['osm_id', 'intensity'])

            for i in range(0, N_MUN):
                try:
                    mun = 'Нижегородская ' + municipalities[i]
                    mun = mun.replace('г.', '')
                    osmnx_row = ox.geocode_to_gdf(mun)

                except ValueError:
                    first_index = municipalities[i].find(' ')
                    second_index = municipalities[i].rindex(' ')
                    mun = 'Нижегородская ' + municipalities[i][:first_index] + municipalities[i][second_index:]
                    osmnx_row = ox.geocode_to_gdf(mun)

                prepare_data()

            df_all = pd.merge(df_all, id2intensity, on='osm_id')
            df_all = geopandas.GeoDataFrame(df_all)

            ch = folium.features.Choropleth(
                geo_data=df_all,
                data=id2intensity,
                columns=['osm_id', 'intensity'],
                key_on='feature.properties.osm_id',
                fill_color="YlOrRd",
                fill_opacity=0.9,
                legend_name=values['par2']
            ).add_to(m)

            ch.geojson.add_child(folium.features.GeoJsonTooltip(aliases=['Название:', 'Значение:'],
                                                                fields=['display_name', 'intensity']))

            template = f"""
            {{% macro html(this, kwargs) %}}
            <!doctype html>
            <html lang="en">
            <head>
              <meta charset="utf-8">
              <meta name="viewport" content="width=device-width, initial-scale=1">
            </head>
            <body>
            <div id='maplegend' class='maplegend'>
              <div class='legend-scale'>
                <ul class='legend-labels'>
            """

            for label in legend:
                template += f"<li>{label}</li>"

            template += """
                </ul>
              </div>
            </div>
            </body>
            </html>

            <style type='text/css'>
              .maplegend {
                position: absolute;
                z-index:9999;
                background-color: rgba(255, 255, 255, 1);
                border-radius: 5px;
                border: 2px solid #bbb;
                padding: 10px;
                font-size:12px;
                left: 0px;
                top: 0px;
              }
              .maplegend .legend-scale ul {
                margin: 0;
                margin-bottom: 5px;
                padding: 0;
                float: left;
                list-style: none;
                }
              .maplegend .legend-scale ul li {
                font-size: 80%;
                list-style: none;
                margin-left: 0;
                line-height: 14px;
                margin-bottom: 2px;
                }
            </style>
            {% endmacro %}
            """

            macro = MacroElement()
            macro._template = Template(template)
            m.get_root().add_child(macro)

            folium.FitBounds([[NO.bounds.miny[0], NO.bounds.minx[0]], [NO.bounds.maxy[0], NO.bounds.maxx[0]]]).add_to(m)

            m.save(html_page)
            webbrowser.open('file://' + os.path.realpath(html_page))

window.close()
